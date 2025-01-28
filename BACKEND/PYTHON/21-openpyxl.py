from openpyxl import load_workbook , Workbook

import datetime
import os 

file_path = os.path.join("C:\\github\\joaoG23\\pessoal\\Novo-repositorio-estudos\\BACKEND\\PYTHON", 'example.xlsx')

def criar_planilha():
    wb = Workbook()

    ws = wb.active

    ws['A1'] = 42

    # Adicionar linha abaixo da outra
    ws.append([1, 2, 3])
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    # ws['A2'] = datetime.datetime.now()


    wb.save(file_path)

# criar_planilha()

def abrir_planilha():
    # Carrega planinha
    ws = load_workbook(file_path)
    ws.active
    
    # print(ws.sheetnames) # ['Sheet', 'Planilha1']
    
    aba_planilha1 = ws['Planilha1']
    # print(aba_planilha1) # seleciona a aba
    celulaA1 = aba_planilha1['A1']
    print(celulaA1.value) # Valor da celula A1
    valor_a1_modo_cell = aba_planilha1.cell(row=1, column=1)
    print(valor_a1_modo_cell.value) # Valor da celula a1

def atualizar_planilha():
    wb = load_workbook(file_path)
    
    aba_planilha1 = wb.active  # Já pega a primeira aba ativa
    
    celulaA1 = aba_planilha1['A1']
    
    celulaA1.value = 'Novo valor'
    
    valor_a2_modo_cell = aba_planilha1.cell(row=1, column=2)
    valor_a2_modo_cell.value = '2012'
    
    wb.save(file_path)
    


def buscar_ultima_linha():
    wb = load_workbook(file_path)
    
    aba_planilha1 = wb['Planilha1']

    # print(aba_planilha1.max_row)
    print(len(aba_planilha1['A'])) # Melhor opção
    
   
def criar_aba():
    wb = load_workbook(file_path)
    

    wb.create_sheet('Pas')
    wb.save(file_path)
    
if __name__ == '__main__':
    criar_aba()
    # buscar_ultima_linha()